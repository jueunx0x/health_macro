import selenium
from selenium import webdriver
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
import time

from selenium.webdriver.support.expected_conditions import presence_of_element_located
from selenium.webdriver.support.wait import WebDriverWait


import openpyxl
from openpyxl import load_workbook

filename = "C:/Users/user/Desktop/list.xlsx"
load_wb = openpyxl.load_workbook(filename,data_only=True)
load_ws=load_wb['Sheet1']
a=[]

import selenium
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
import time

from selenium.webdriver.support.expected_conditions import presence_of_element_located
from selenium.webdriver.support.wait import WebDriverWait

url = ""#접속할 사이트

options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])
#driver = webdriver.Chrome(options=options)
driver=webdriver.Chrome(ChromeDriverManager().install())
driver.maximize_window()
driver.get(url)  # driver실행

# 로그인
time.sleep(5)
driver.switch_to.frame('base')

#공인인증서
time.sleep(5)





driver.switch_to.frame('base')
driver.find_element(By.XPATH, '//*[@id="mCSB_1_container"]/ul/li[3]/a').click()
time.sleep(3)
driver.find_element(By.XPATH,'//*[@id="mCSB_1_container"]/ul/li[3]/ul/li[1]/a').click()
driver.switch_to.frame('ifrm')
for i in range(2,52):

    time.sleep(3)
    name=load_ws.cell(i,3).value
    if name=='None':
        break

    print(name)

    social_num=load_ws.cell(i,5).value#.value.replace("-","")
    time.sleep(3)
    driver.find_element(By.XPATH, '//*[@id="mtxtPatntNm"]').click()
    elements=driver.find_element(By.XPATH, '//*[@id="mtxtPatntNm"]')

    #이름 검색
    elements.clear()
    elements.send_keys(name)
    driver.find_element(By.XPATH, '//*[@id="mtxtPatntNm"]').click()
    time.sleep(3)
    #주민번호검색
    elements = driver.find_element(By.XPATH, '//*[@id="mtxtPatntIhidnum"]')
    elements.clear()
    elements.send_keys(social_num)
    driver.find_element(By.XPATH,'//*[@id="mbtnSearch"]').click()
    time.sleep(2)
    driver.find_element(By.XPATH, '// *[ @ id = "jqg_grdSelfSmsSendRequst_1"]').click()


    time.sleep(3)
    elements = driver.find_element(By.XPATH, '//*[@id="smsSendRequstBtn"]').click()

    time.sleep(1)
    da = Alert(driver)
    da.accept()
    time.sleep(3)
    '''
    time.sleep(3)
    driver.find_element(By.XPATH,'//*[@id="jqg_grdSelfSmsSendRequst_1"]').click()
    driver.find_element(By.XPATH,'//*[@id="smsSendRequstBtn"]').click()'''

    driver.find_element(By.XPATH,'//*[@id="smsSendBtn"]').click()

    time.sleep(3)
    da = Alert(driver)
    da.accept()

driver.quit()