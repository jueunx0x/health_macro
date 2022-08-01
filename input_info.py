import selenium
import urllib3
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
import time

from selenium.webdriver.support.expected_conditions import presence_of_element_located
from selenium.webdriver.support.wait import WebDriverWait

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string

from datetime import datetime,timedelta

from webdriver_manager.chrome import ChromeDriverManager


def check_exists_by_XPATH(xpath):
    try:
        driver.find_element(By.XPATH,xpath)
    except NoSuchElementException:
        return False
    return True

url = ""#접속할 사이트

options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])
#driver = webdriver.Chrome(options=options)
driver=webdriver.Chrome(ChromeDriverManager().install())
driver.maximize_window()

driver.get(url)  # driver실행

# 로그인
time.sleep(7)
driver.switch_to.frame('base')
#공인인증서
# 역학조사
time.sleep(5)
driver.switch_to.frame('base')
driver.find_element(By.XPATH, '//*[@id="mCSB_1_container"]/ul/li[3]/a').click()
time.sleep(5)
driver.find_element(By.XPATH, '//*[@id="mCSB_1_container"]/ul/li[3]/ul/li[2]/a').click()

time.sleep(20)
#날짜 및 임시저장 전부 선택하기
element = driver.switch_to.frame("ifrm")  # 프레임 변경


#반복문 시작:
while True:
    try:
        exist = check_exists_by_XPATH('//*[@id="contetnsWrap"]/tbody/tr/td/div[3]/div[2]/button[1]')
        if(exist==True):
            driver.find_element(By.XPATH,
                                '//*[@id="contetnsWrap"]/tbody/tr/td/div[3]/div[2]/button[1]').click()  # 조회버튼 클릭
            time.sleep(5)
            driver.find_element(By.XPATH, '//*[@id="1"]/td[1]').click()
            driver.find_element(By.XPATH, '//*[@id="contetnsWrap"]/tbody/tr/td/div[3]/div[2]/button[4]').click() #상세보기
            time.sleep(4)
            #driver.find_element(By.XPATH,
            #                    '//*[@id="regist_pop"]/div/div[1]/div[1]/table/tbody/tr/td/table/tbody/tr[3]/td[2]/img').click()  # 날짜선택
            #driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[2]/a').click()  # 오늘날짜로 선택

            # 주민번호
            number1 = str(driver.find_element(By.XPATH, '//*[@id="patntIhidnum1"]').get_attribute("value"))
            number2 = str(driver.find_element(By.XPATH, '//*[@id="patntIhidnum2"]').get_attribute("value"))
            total_num = str(number1 + '-' + number2)

            filename = "C:/Users/user/Desktop/day.xlsx"
            load_wb = openpyxl.load_workbook(filename, data_only=True)
            load_ws = load_wb['Sheet1']

            for i in range(1, load_ws.max_row):
                if str(load_ws.cell(i, 2).value) == total_num:
                    a = load_ws.cell(i, 2).row
                    day = load_ws.cell(a, 6).value
                    day = datetime.strptime(day, '%Y%m%d')
                    year = day.year
                    month = day.month
                    date = day.day
                    driver.find_element(By.XPATH,
                                        '//*[@id="regist_pop"]/div/div[1]/div[3]/table/tbody/tr/td/table/tbody/tr[2]/td/img').click()

                    # 1주차
                    if date == 1:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[6]/a').click()
                    elif date == 2:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[7]/a').click()
                    # 2주차
                    elif date == 3:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[1]/a').click()
                    elif date == 4:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[2]/a').click()
                    elif date == 5:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[3]/a').click()
                    elif date == 6:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[4]/a').click()
                    elif date == 7:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[5]/a').click()
                    elif date == 8:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[6]/a').click()
                    elif date == 9:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[7]/a').click()
                    # 3주차
                    elif date == 10:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[1]/a').click()
                    elif date == 11:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[2]/a').click()
                    elif date == 12:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[3]/a').click()
                    elif date == 13:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[4]/a').click()
                    elif date == 14:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[5]/a').click()
                    elif date == 15:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[6]/a').click()
                    elif date == 16:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[7]/a').click()
                    # 4주차
                    elif date == 17:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[1]/a').click()
                    elif date == 18:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[2]/a').click()
                    elif date == 19:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[3]/a').click()
                    elif date == 20:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[4]/a').click()
                    elif date == 21:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[5]/a').click()
                    elif date == 22:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[6]/a').click()
                    elif date == 23:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[7]/a').click()
                    # 5주차
                    elif date == 24:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[1]/a').click()
                    elif date == 25:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()
                    elif date == 26:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[3]/a').click()
                    elif date == 27:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[4]/a').click()
                    elif date == 28:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[5]/a').click()
                    elif date == 29:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()
                    elif date == 30:
                        driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[7]/a').click()
                    else:
                        print(date, "없음")
            # 본인여부
            self = driver.find_element(By.XPATH, '//*[@id="patntSelfSe"and @value="1"]')
            parent = driver.find_element(By.XPATH, '//*[@id="patntSelfSe" and @value="2"]')

            if self.is_selected() == False and parent.is_selected() == False:
                self.click()

            # 장기입원환자 여부
            hospital_y = driver.find_element(By.XPATH, '//*[@id="constntAt"and @value="Y"]')
            hospital_N = driver.find_element(By.XPATH, '//*[@id="constntAt"and @value="N"]')
            if hospital_y.is_selected() == False and hospital_N.is_selected() == False:
                hospital_N.click()

            # 증상여부
            symptom_radio_y = driver.find_element(By.XPATH, '//*[@id="smptmAt"and @value="1"]')
            symptom_radio_N = driver.find_element(By.XPATH, '//*[@id="smptmAt"and @value="2"]')

            if symptom_radio_y.is_selected() == False and symptom_radio_N.is_selected() == False:
                symptom_radio_N.click()
            # 백신여부
            vaccine_radio_Y = driver.find_element(By.XPATH, '//*[@id="covidVacntnAt"and @value="1"]')
            vaccine_radio_N = driver.find_element(By.XPATH, '//*[@id="covidVacntnAt"and @value="2"]')

            if vaccine_radio_Y.is_selected() == False and vaccine_radio_N.is_selected() == False:
                vaccine_radio_N.click()
            # 가족 여부
            family_button_Y = driver.find_element(By.XPATH, '//*[@id="prnctFmlyRmt"and @value="1"]')
            family_button_N = driver.find_element(By.XPATH, '//*[@id="prnctFmlyRmt"and @value="2"]')

            if family_button_Y.is_selected() == False and family_button_N.is_selected() == False:
                family_button_N.click()

            driver.find_element(By.XPATH, '//*[@id="mrfnSecrtryNtcnSe" and @value="9"]').click()  # 국민비서

            driver.find_element(By.XPATH, '//*[@id="manageMthdSe" and @value="1"]').click()  # 재택치료

            driver.find_element(By.XPATH, '//*[@id="btn_modi"]').click()  # 저장

            time.sleep(3)
            da = Alert(driver)
            da.accept()
            time.sleep(5)
            da.accept()
            time.sleep(4)
        else:
            driver.quit()
    except urllib3.exceptions.MaxRetryError:
        driver.quit()

'''
    
time.sleep(5)
driver.find_element(By.XPATH,'//*[@id="1"]/td[1]').click()
driver.find_element(By.XPATH,'//*[@id="contetnsWrap"]/tbody/tr/td/div[3]/div[2]/button[4]').click()
time.sleep(4)
driver.find_element(By.XPATH,'//*[@id="regist_pop"]/div/div[1]/div[1]/table/tbody/tr/td/table/tbody/tr[3]/td[2]/img').click()#날짜선택
driver.find_element(By.XPATH,'//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[4]/a').click()#오늘날짜로 선택

#주민번호
number1=str(driver.find_element(By.XPATH,'//*[@id="patntIhidnum1"]').get_attribute("value"))
number2=str(driver.find_element(By.XPATH,'//*[@id="patntIhidnum2"]').get_attribute("value"))
total_num=str(number1+'-'+number2)

filename = "C:/Users/user/Downloads/day.xlsx"
load_wb = openpyxl.load_workbook(filename,data_only=True)
load_ws=load_wb['Sheet1']

for i in range(1, load_ws.max_row):
    if str(load_ws.cell(i, 2).value) == total_num:
        a = load_ws.cell(i, 2).row
        day = load_ws.cell(a, 6).value
        day = datetime.strptime(day, '%Y%m%d')
        year = day.year
        month = day.month
        date = day.day
        driver.find_element(By.XPATH,'//*[@id="regist_pop"]/div/div[1]/div[3]/table/tbody/tr/td/table/tbody/tr[2]/td/img').click()

        # 1주차
        if date == 1:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[6]/a').click()
        elif date == 2:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[7]/a').click()
        # 2주차
        elif date == 3:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[1]/a').click()
        elif date == 4:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[2]/a').click()
        elif date == 5:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[3]/a').click()
        elif date == 6:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[4]/a').click()
        elif date == 7:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[5]/a').click()
        elif date == 8:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[6]/a').click()
        elif date == 9:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[7]/a').click()
        # 3주차
        elif date == 10:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[1]/a').click()
        elif date == 11:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[2]/a').click()
        elif date == 12:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[3]/a').click()
        elif date == 13:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[4]/a').click()
        elif date == 14:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[5]/a').click()
        elif date == 15:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[6]/a').click()
        elif date == 16:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[7]/a').click()
        # 4주차
        elif date == 17:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[1]/a').click()
        elif date == 18:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[2]/a').click()
        elif date == 19:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[3]/a').click()
        elif date == 20:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[4]/a').click()
        elif date == 21:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[5]/a').click()
        elif date == 22:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[6]/a').click()
        elif date == 23:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[7]/a').click()
        # 5주차
        elif date == 24:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[1]/a').click()
        elif date == 25:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()
        elif date == 26:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[3]/a').click()
        elif date == 27:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[4]/a').click()
        elif date == 28:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[5]/a').click()
        elif date == 29:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()
        elif date == 30:
            driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[7]/a').click()
        else:
            print(date, "없음")
#본인여부
self = driver.find_element(By.XPATH, '//*[@id="patntSelfSe"and @value="1"]')
parent = driver.find_element(By.XPATH, '//*[@id="patntSelfSe" and @value="2"]')


if self.is_selected() == False and parent.is_selected() == False:
    self.click()

#장기입원환자 여부
hospital_y=driver.find_element(By.XPATH,'//*[@id="constntAt"and @value="Y"]')
hospital_N=driver.find_element(By.XPATH,'//*[@id="constntAt"and @value="N"]')
if hospital_y.is_selected()==False and hospital_N.is_selected()==False:
    hospital_N.click()

#증상여부
symptom_radio_y = driver.find_element(By.XPATH, '//*[@id="smptmAt"and @value="1"]')
symptom_radio_N = driver.find_element(By.XPATH, '//*[@id="smptmAt"and @value="2"]')

if symptom_radio_y.is_selected() == False and symptom_radio_N.is_selected() == False:
    symptom_radio_N.click()
#백신여부
vaccine_radio_Y = driver.find_element(By.XPATH, '//*[@id="covidVacntnAt"and @value="1"]')
vaccine_radio_N = driver.find_element(By.XPATH, '//*[@id="covidVacntnAt"and @value="2"]')

if vaccine_radio_Y.is_selected() == False and vaccine_radio_N.is_selected() == False:
    vaccine_radio_N.click()
#가족 여부
family_button_Y = driver.find_element(By.XPATH, '//*[@id="prnctFmlyRmt"and @value="1"]')
family_button_N = driver.find_element(By.XPATH, '//*[@id="prnctFmlyRmt"and @value="2"]')

if family_button_Y.is_selected() == False and family_button_N.is_selected() == False:
    family_button_N.click()

driver.find_element(By.XPATH, '//*[@id="mrfnSecrtryNtcnSe" and @value="9"]').click()  # 국민비서

driver.find_element(By.XPATH, '//*[@id="manageMthdSe" and @value="1"]').click()  # 재택치료


    driver.find_element(By.XPATH,'//*[@id="btn_modi"]').click()#저장

    time.sleep(3)
    da = Alert(driver)
    da.accept()
    time.sleep(5)
    da.accept()
    time.sleep(4)

'''

#반복문 종료.