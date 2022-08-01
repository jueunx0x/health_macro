import selenium
from selenium import webdriver
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
import time

from selenium.webdriver.support.expected_conditions import presence_of_element_located
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

import openpyxl
from openpyxl import load_workbook

filename = "C:/Users/user/Desktop/list.xlsx"
load_wb = openpyxl.load_workbook(filename,data_only=True)
load_ws=load_wb['Sheet1']

import selenium
from selenium import webdriver
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
import time

from selenium.webdriver.support.expected_conditions import presence_of_element_located
from selenium.webdriver.support.wait import WebDriverWait

url = ""

options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])
#driver = webdriver.Chrome(options=options)
driver=webdriver.Chrome(ChromeDriverManager().install())
driver.maximize_window()

driver.get(url)  # driver실행

# 로그인
time.sleep(20)

#공인인증서

#웹보고 들어가기
time.sleep(5)
driver.switch_to.frame('base')
driver.find_element(By.XPATH, '//*[@id="mCSB_1_container"]/ul/li[1]/a').click()
time.sleep(2)
driver.find_element(By.XPATH, '//*[@id="mCSB_1_container"]/ul/li[1]/ul/li/a').click()
time.sleep(2)
driver.find_element(By.XPATH, '//*[@id="mCSB_1_container"]/ul/li[1]/ul/li/ul/li[1]/a').click()
driver.switch_to.frame('ifrm')



for i in range(2,101):
    time.sleep(3)
    name=load_ws.cell(i,3).value
    if name=='None':
        break
    social_num=load_ws.cell(i,5).value
    time.sleep(3)
    driver.find_element(By.XPATH, '//*[@id="mtxtPatntNm"]').click()
    elements=driver.find_element(By.XPATH, '//*[@id="mtxtPatntNm"]')

    #이름 검색
    elements.clear()
    elements.send_keys(name)

    #주민번호검색
    driver.find_element(By.XPATH, '//*[@id="mtxtPatntIhidnum"]').click()
    elements = driver.find_element(By.XPATH, '//*[@id="mtxtPatntIhidnum"]')
    elements.clear()
    elements.send_keys(social_num)

    time.sleep(3)
    elements = driver.find_element(By.XPATH, '//*[@id="mbtnSearch"]').click()

    table = driver.find_element(By.XPATH, '//*[@id="grdInfectionsReportList"]')
    tbody = table.find_element(By.TAG_NAME, "tbody")
    rows = tbody.find_element(By.TAG_NAME, "tr")

    time.sleep(3)

    driver.find_element(By.XPATH, '//*[@id="1"]/td[14]').click()

    time.sleep(4)
    driver.find_element(By.XPATH, '//*[@id="subDetailForm"]/div/div/button[1]').click()
    time.sleep(2)

    da = Alert(driver)
    da.accept()
    time.sleep(4)
    driver.find_element(By.XPATH, '//*[@id="pbtnCreateReport"]').click()
    time.sleep(2)
    da.accept()
